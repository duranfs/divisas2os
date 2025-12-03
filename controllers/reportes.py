# -*- coding: utf-8 -*-
"""
Controlador de Reportes - Sistema de Divisas Bancario
Maneja la generación de reportes y consultas de historial
Requisitos: 6.1, 6.2, 6.4, 7.1, 7.3
"""

import logging
import datetime
import json
from decimal import Decimal
from io import BytesIO
import os

# Configurar logging
logger = logging.getLogger("web2py.app.reportes")

def obtener_cuentas_transaccion(transaccion):
    """
    Función auxiliar para obtener las cuentas origen y destino de una transacción
    basándose en el modelo actual donde cada cliente tiene cuentas por moneda.
    
    Args:
        transaccion: Row de transacción con cuenta_id, tipo_operacion, moneda_origen, moneda_destino
    
    Returns:
        tuple: (cuenta_origen_str, cuenta_destino_str)
    """
    try:
        # Obtener la cuenta principal de la transacción
        cuenta = db(db.cuentas.id == transaccion.cuenta_id).select().first() if transaccion.cuenta_id else None
        
        if not cuenta:
            return ("N/A", "N/A")
        
        cliente_id = cuenta.cliente_id
        
        if transaccion.tipo_operacion == 'compra':
            # En compra: cuenta_id es VES (origen), buscar cuenta destino en la moneda comprada
            cuenta_origen_str = f"{cuenta.numero_cuenta[-4:]}(VES)"
            cuenta_destino = db(
                (db.cuentas.cliente_id == cliente_id) &
                (db.cuentas.moneda == transaccion.moneda_destino)
            ).select().first()
            cuenta_destino_str = f"{cuenta_destino.numero_cuenta[-4:]}({transaccion.moneda_destino})" if cuenta_destino else f"N/A({transaccion.moneda_destino})"
            
        else:  # venta
            # En venta: cuenta_id es divisa (origen), buscar cuenta VES (destino)
            cuenta_origen_str = f"{cuenta.numero_cuenta[-4:]}({transaccion.moneda_origen})"
            cuenta_destino = db(
                (db.cuentas.cliente_id == cliente_id) &
                (db.cuentas.moneda == 'VES')
            ).select().first()
            cuenta_destino_str = f"{cuenta_destino.numero_cuenta[-4:]}(VES)" if cuenta_destino else "N/A(VES)"
        
        return (cuenta_origen_str, cuenta_destino_str)
        
    except Exception as e:
        logger.error(f"Error en obtener_cuentas_transaccion: {str(e)}")
        return ("N/A", "N/A")

def index():
    """
    Página principal del módulo de reportes
    Muestra opciones de consulta y generación de reportes
    """
    try:
        # Verificar permisos de usuario
        if not auth.user:
            redirect(URL('default', 'user/login'))
        
        # Obtener estadísticas básicas para el dashboard
        estadisticas = obtener_estadisticas_generales()
        
        return dict(
            estadisticas=estadisticas,
            mensaje="Módulo de Reportes y Consultas"
        )
        
    except Exception as e:
        logger.error(f"Error en index de reportes: {str(e)}")
        response.flash = f"Error cargando módulo de reportes: {str(e)}"
        return dict(
            estadisticas=None,
            mensaje="Error en el sistema"
        )

@auth.requires_login()
def historial_transacciones():
    """
    Función de historial de transacciones con filtros
    Requisitos: 6.1, 6.2
    """
    try:
        # Verificar si es administrador o cliente
        es_admin = auth.has_membership('administrador')
        
        if not es_admin:
            # Para clientes, mostrar solo sus transacciones
            cliente = db(db.clientes.user_id == auth.user.id).select().first()
            if not cliente:
                response.flash = "Debe completar su registro como cliente"
                redirect(URL('clientes', 'registrar'))
        
        # Obtener parámetros de filtro
        fecha_desde = request.vars.fecha_desde
        fecha_hasta = request.vars.fecha_hasta
        tipo_operacion = request.vars.tipo_operacion
        moneda = request.vars.moneda
        cliente_id = request.vars.cliente_id if es_admin else None
        
        # Construir query base
        if es_admin:
            # Administradores ven todas las transacciones
            query = (db.transacciones.cuenta_id == db.cuentas.id) & \
                    (db.cuentas.cliente_id == db.clientes.id) & \
                    (db.clientes.user_id == db.auth_user.id)
        else:
            # Clientes ven solo sus transacciones
            query = (db.transacciones.cuenta_id == db.cuentas.id) & \
                    (db.cuentas.cliente_id == cliente.id)
        
        # Aplicar filtros
        if fecha_desde:
            try:
                fecha_desde_dt = datetime.datetime.strptime(fecha_desde, '%Y-%m-%d')
                query &= db.transacciones.fecha_transaccion >= fecha_desde_dt
            except ValueError:
                response.flash = "Formato de fecha desde inválido"
        
        if fecha_hasta:
            try:
                fecha_hasta_dt = datetime.datetime.strptime(fecha_hasta, '%Y-%m-%d')
                fecha_hasta_dt = fecha_hasta_dt.replace(hour=23, minute=59, second=59)
                query &= db.transacciones.fecha_transaccion <= fecha_hasta_dt
            except ValueError:
                response.flash = "Formato de fecha hasta inválido"
        
        if tipo_operacion and tipo_operacion != 'todos':
            query &= db.transacciones.tipo_operacion == tipo_operacion
        
        if moneda and moneda != 'todas':
            query &= ((db.transacciones.moneda_origen == moneda) | 
                     (db.transacciones.moneda_destino == moneda))
        
        if cliente_id and es_admin:
            query &= db.clientes.id == cliente_id
        
        # Obtener transacciones con información relacionada
        if es_admin:
            transacciones = db(query).select(
                db.transacciones.ALL,
                db.cuentas.numero_cuenta,
                db.clientes.cedula,
                db.auth_user.first_name,
                db.auth_user.last_name,
                join=[
                    db.cuentas.on(db.transacciones.cuenta_id == db.cuentas.id),
                    db.clientes.on(db.cuentas.cliente_id == db.clientes.id),
                    db.auth_user.on(db.clientes.user_id == db.auth_user.id)
                ],
                orderby=~db.transacciones.fecha_transaccion,
                limitby=(0, 100)  # Limitar a 100 registros por página
            )
        else:
            transacciones = db(query).select(
                db.transacciones.ALL,
                db.cuentas.numero_cuenta,
                join=db.cuentas.on(db.transacciones.cuenta_id == db.cuentas.id),
                orderby=~db.transacciones.fecha_transaccion,
                limitby=(0, 100)
            )
        
        # Obtener lista de clientes para filtro (solo admin)
        clientes_lista = []
        if es_admin:
            clientes_lista = db(
                db.clientes.id == db.auth_user.id
            ).select(
                db.clientes.id,
                db.clientes.cedula,
                db.auth_user.first_name,
                db.auth_user.last_name,
                join=db.auth_user.on(db.clientes.user_id == db.auth_user.id),
                orderby=db.auth_user.first_name
            )
        
        return dict(
            transacciones=transacciones,
            es_admin=es_admin,
            clientes_lista=clientes_lista,
            filtros={
                'fecha_desde': fecha_desde,
                'fecha_hasta': fecha_hasta,
                'tipo_operacion': tipo_operacion,
                'moneda': moneda,
                'cliente_id': cliente_id
            }
        )
        
    except Exception as e:
        logger.error(f"Error en historial_transacciones: {str(e)}")
        response.flash = f"Error: {str(e)}"
        redirect(URL('reportes', 'index'))

@auth.requires_membership('administrador')
def reportes_administrativos():
    """
    Implementar reportes administrativos diarios
    Requisitos: 7.1, 7.3, 8.1, 8.2, 8.3
    """
    try:
        # Obtener parámetros
        fecha_reporte = request.vars.fecha_reporte
        tipo_reporte = request.vars.tipo_reporte or 'diario'
        moneda_filtro = request.vars.moneda_filtro or 'todas'  # Nuevo: filtro por moneda
        
        if not fecha_reporte:
            fecha_reporte = datetime.date.today().strftime('%Y-%m-%d')
        
        # Generar reporte según el tipo
        if tipo_reporte == 'diario':
            reporte = generar_reporte_diario(fecha_reporte, moneda_filtro)
        elif tipo_reporte == 'semanal':
            reporte = generar_reporte_semanal(fecha_reporte, moneda_filtro)
        elif tipo_reporte == 'mensual':
            reporte = generar_reporte_mensual(fecha_reporte, moneda_filtro)
        else:
            reporte = {'error': 'Tipo de reporte no válido'}
        
        return dict(
            reporte=reporte,
            fecha_reporte=fecha_reporte,
            tipo_reporte=tipo_reporte,
            moneda_filtro=moneda_filtro
        )
        
    except Exception as e:
        logger.error(f"Error en reportes_administrativos: {str(e)}")
        response.flash = f"Error: {str(e)}"
        return dict(
            reporte={'error': str(e)},
            fecha_reporte=fecha_reporte,
            tipo_reporte=tipo_reporte,
            moneda_filtro='todas'
        )

def generar_reporte_diario(fecha_str, moneda_filtro='todas'):
    """
    Genera reporte diario de transacciones con soporte para cuentas por moneda
    Requisitos: 8.1, 8.2, 8.3
    """
    try:
        fecha = datetime.datetime.strptime(fecha_str, '%Y-%m-%d').date()
        fecha_inicio = datetime.datetime.combine(fecha, datetime.time.min)
        fecha_fin = datetime.datetime.combine(fecha, datetime.time.max)
        
        # Obtener transacciones del día con filtro de moneda
        query = (db.transacciones.fecha_transaccion >= fecha_inicio) & \
                (db.transacciones.fecha_transaccion <= fecha_fin)
        
        # Aplicar filtro por moneda si se especifica
        if moneda_filtro and moneda_filtro != 'todas':
            query &= ((db.transacciones.moneda_origen == moneda_filtro) | 
                     (db.transacciones.moneda_destino == moneda_filtro))
        
        transacciones = db(query).select()
        
        # Calcular estadísticas
        total_transacciones = len(transacciones)
        compras = [t for t in transacciones if t.tipo_operacion == 'compra']
        ventas = [t for t in transacciones if t.tipo_operacion == 'venta']
        
        # Volúmenes de compras por moneda (lo que se pagó en VES y lo que se recibió en divisas)
        volumen_compras_ves = sum([float(t.monto_origen) for t in compras if t.moneda_origen == 'VES'])
        volumen_compras_usd = sum([float(t.monto_destino) for t in compras if t.moneda_destino == 'USD'])
        volumen_compras_usdt = sum([float(t.monto_destino) for t in compras if t.moneda_destino == 'USDT'])
        volumen_compras_eur = sum([float(t.monto_destino) for t in compras if t.moneda_destino == 'EUR'])
        
        # Volúmenes de ventas por moneda (lo que se entregó en divisas y lo que se recibió en VES)
        volumen_ventas_usd = sum([float(t.monto_origen) for t in ventas if t.moneda_origen == 'USD'])
        volumen_ventas_usdt = sum([float(t.monto_origen) for t in ventas if t.moneda_origen == 'USDT'])
        volumen_ventas_eur = sum([float(t.monto_origen) for t in ventas if t.moneda_origen == 'EUR'])
        volumen_ventas_ves = sum([float(t.monto_destino) for t in ventas if t.moneda_destino == 'VES'])
        
        total_comisiones = sum([float(t.comision) for t in transacciones])
        
        # Calcular tasas promedio por moneda desde las transacciones del día
        transacciones_usd = [t for t in transacciones if t.moneda_origen == 'USD' or t.moneda_destino == 'USD']
        transacciones_usdt = [t for t in transacciones if t.moneda_origen == 'USDT' or t.moneda_destino == 'USDT']
        transacciones_eur = [t for t in transacciones if t.moneda_origen == 'EUR' or t.moneda_destino == 'EUR']
        
        tasa_usd_promedio = sum([float(t.tasa_aplicada) for t in transacciones_usd]) / len(transacciones_usd) if transacciones_usd else 0
        tasa_usdt_promedio = sum([float(t.tasa_aplicada) for t in transacciones_usdt]) / len(transacciones_usdt) if transacciones_usdt else 0
        tasa_eur_promedio = sum([float(t.tasa_aplicada) for t in transacciones_eur]) / len(transacciones_eur) if transacciones_eur else 0
        
        # Obtener tasas actuales para conversión consolidada
        tasas_actuales = obtener_tasas_actuales()
        
        # Calcular equivalente total en VES (consolidado)
        total_consolidado_ves = volumen_compras_ves + volumen_ventas_ves
        if tasas_actuales:
            total_consolidado_ves += (volumen_compras_usd + volumen_ventas_usd) * tasas_actuales.get('USD', 0)
            total_consolidado_ves += (volumen_compras_usdt + volumen_ventas_usdt) * tasas_actuales.get('USDT', 0)
            total_consolidado_ves += (volumen_compras_eur + volumen_ventas_eur) * tasas_actuales.get('EUR', 0)
        
        return {
            'fecha': fecha_str,
            'moneda_filtro': moneda_filtro,
            'total_transacciones': total_transacciones,
            'total_compras': len(compras),
            'total_ventas': len(ventas),
            'volumen_compras_ves': volumen_compras_ves,
            'volumen_compras_usd': volumen_compras_usd,
            'volumen_compras_usdt': volumen_compras_usdt,
            'volumen_compras_eur': volumen_compras_eur,
            'volumen_ventas_usd': volumen_ventas_usd,
            'volumen_ventas_usdt': volumen_ventas_usdt,
            'volumen_ventas_eur': volumen_ventas_eur,
            'volumen_ventas_ves': volumen_ventas_ves,
            'total_comisiones': total_comisiones,
            'tasa_usd_promedio': round(tasa_usd_promedio, 4),
            'tasa_usdt_promedio': round(tasa_usdt_promedio, 4),
            'tasa_eur_promedio': round(tasa_eur_promedio, 4),
            'total_consolidado_ves': round(total_consolidado_ves, 2),
            'tasas_conversion': tasas_actuales,
            'transacciones_detalle': [
                {
                    'id': t.id,
                    'comprobante': t.numero_comprobante,
                    'tipo': t.tipo_operacion,
                    'cuenta_id': t.cuenta_id,
                    'monto_origen': float(t.monto_origen),
                    'moneda_origen': t.moneda_origen,
                    'monto_destino': float(t.monto_destino),
                    'moneda_destino': t.moneda_destino,
                    'tasa': float(t.tasa_aplicada),
                    'comision': float(t.comision),
                    'fecha': t.fecha_transaccion.strftime('%H:%M:%S')
                } for t in transacciones
            ]
        }
        
    except Exception as e:
        logger.error(f"Error generando reporte diario: {str(e)}")
        return {'error': str(e)}

def generar_reporte_semanal(fecha_str, moneda_filtro='todas'):
    """
    Genera reporte semanal de transacciones con soporte para cuentas por moneda
    Requisitos: 8.1, 8.2, 8.3
    """
    try:
        fecha = datetime.datetime.strptime(fecha_str, '%Y-%m-%d').date()
        # Calcular inicio de semana (lunes)
        dias_desde_lunes = fecha.weekday()
        fecha_inicio_semana = fecha - datetime.timedelta(days=dias_desde_lunes)
        fecha_fin_semana = fecha_inicio_semana + datetime.timedelta(days=6)
        
        fecha_inicio = datetime.datetime.combine(fecha_inicio_semana, datetime.time.min)
        fecha_fin = datetime.datetime.combine(fecha_fin_semana, datetime.time.max)
        
        # Obtener transacciones de la semana con filtro de moneda
        query = (db.transacciones.fecha_transaccion >= fecha_inicio) & \
                (db.transacciones.fecha_transaccion <= fecha_fin)
        
        if moneda_filtro and moneda_filtro != 'todas':
            query &= ((db.transacciones.moneda_origen == moneda_filtro) | 
                     (db.transacciones.moneda_destino == moneda_filtro))
        
        transacciones = db(query).select()
        
        # Agrupar por día
        transacciones_por_dia = {}
        for t in transacciones:
            dia = t.fecha_transaccion.date()
            if dia not in transacciones_por_dia:
                transacciones_por_dia[dia] = []
            transacciones_por_dia[dia].append(t)
        
        # Obtener tasas actuales para conversión
        tasas_actuales = obtener_tasas_actuales()
        
        # Calcular estadísticas por día
        resumen_diario = []
        for i in range(7):
            dia = fecha_inicio_semana + datetime.timedelta(days=i)
            transacciones_dia = transacciones_por_dia.get(dia, [])
            
            compras_dia = [t for t in transacciones_dia if t.tipo_operacion == 'compra']
            ventas_dia = [t for t in transacciones_dia if t.tipo_operacion == 'venta']
            
            # Calcular volúmenes por moneda
            vol_ves = sum([float(t.monto_origen) for t in compras_dia if t.moneda_origen == 'VES']) + \
                      sum([float(t.monto_destino) for t in ventas_dia if t.moneda_destino == 'VES'])
            vol_usd = sum([float(t.monto_destino) for t in compras_dia if t.moneda_destino == 'USD']) + \
                      sum([float(t.monto_origen) for t in ventas_dia if t.moneda_origen == 'USD'])
            vol_usdt = sum([float(t.monto_destino) for t in compras_dia if t.moneda_destino == 'USDT']) + \
                       sum([float(t.monto_origen) for t in ventas_dia if t.moneda_origen == 'USDT'])
            vol_eur = sum([float(t.monto_destino) for t in compras_dia if t.moneda_destino == 'EUR']) + \
                      sum([float(t.monto_origen) for t in ventas_dia if t.moneda_origen == 'EUR'])
            
            # Calcular equivalente en VES
            vol_ves_equiv = vol_ves + \
                           (vol_usd * tasas_actuales.get('USD', 0)) + \
                           (vol_usdt * tasas_actuales.get('USDT', 0)) + \
                           (vol_eur * tasas_actuales.get('EUR', 0))
            
            resumen_diario.append({
                'fecha': dia.strftime('%Y-%m-%d'),
                'dia_semana': dia.strftime('%A'),
                'total_transacciones': len(transacciones_dia),
                'compras': len(compras_dia),
                'ventas': len(ventas_dia),
                'volumen_ves': vol_ves,
                'volumen_usd': vol_usd,
                'volumen_usdt': vol_usdt,
                'volumen_eur': vol_eur,
                'volumen_ves_equivalente': round(vol_ves_equiv, 2),
                'comisiones': sum([float(t.comision) for t in transacciones_dia])
            })
        
        # Totales de la semana
        total_transacciones = len(transacciones)
        total_comisiones = sum([float(t.comision) for t in transacciones])
        
        # Totales por moneda
        compras = [t for t in transacciones if t.tipo_operacion == 'compra']
        ventas = [t for t in transacciones if t.tipo_operacion == 'venta']
        
        total_ves = sum([float(t.monto_origen) for t in compras if t.moneda_origen == 'VES']) + \
                    sum([float(t.monto_destino) for t in ventas if t.moneda_destino == 'VES'])
        total_usd = sum([float(t.monto_destino) for t in compras if t.moneda_destino == 'USD']) + \
                    sum([float(t.monto_origen) for t in ventas if t.moneda_origen == 'USD'])
        total_usdt = sum([float(t.monto_destino) for t in compras if t.moneda_destino == 'USDT']) + \
                     sum([float(t.monto_origen) for t in ventas if t.moneda_origen == 'USDT'])
        total_eur = sum([float(t.monto_destino) for t in compras if t.moneda_destino == 'EUR']) + \
                    sum([float(t.monto_origen) for t in ventas if t.moneda_origen == 'EUR'])
        
        total_consolidado_ves = total_ves + \
                               (total_usd * tasas_actuales.get('USD', 0)) + \
                               (total_usdt * tasas_actuales.get('USDT', 0)) + \
                               (total_eur * tasas_actuales.get('EUR', 0))
        
        return {
            'fecha_inicio': fecha_inicio_semana.strftime('%Y-%m-%d'),
            'fecha_fin': fecha_fin_semana.strftime('%Y-%m-%d'),
            'moneda_filtro': moneda_filtro,
            'total_transacciones': total_transacciones,
            'total_comisiones': total_comisiones,
            'total_ves': total_ves,
            'total_usd': total_usd,
            'total_usdt': total_usdt,
            'total_eur': total_eur,
            'total_consolidado_ves': round(total_consolidado_ves, 2),
            'tasas_conversion': tasas_actuales,
            'resumen_diario': resumen_diario
        }
        
    except Exception as e:
        logger.error(f"Error generando reporte semanal: {str(e)}")
        return {'error': str(e)}

def generar_reporte_mensual(fecha_str, moneda_filtro='todas'):
    """
    Genera reporte mensual de transacciones con soporte para cuentas por moneda
    Requisitos: 8.1, 8.2, 8.3
    """
    try:
        fecha = datetime.datetime.strptime(fecha_str, '%Y-%m-%d').date()
        # Primer y último día del mes
        primer_dia = fecha.replace(day=1)
        if fecha.month == 12:
            ultimo_dia = fecha.replace(year=fecha.year + 1, month=1, day=1) - datetime.timedelta(days=1)
        else:
            ultimo_dia = fecha.replace(month=fecha.month + 1, day=1) - datetime.timedelta(days=1)
        
        fecha_inicio = datetime.datetime.combine(primer_dia, datetime.time.min)
        fecha_fin = datetime.datetime.combine(ultimo_dia, datetime.time.max)
        
        # Obtener transacciones del mes con filtro de moneda
        query = (db.transacciones.fecha_transaccion >= fecha_inicio) & \
                (db.transacciones.fecha_transaccion <= fecha_fin)
        
        if moneda_filtro and moneda_filtro != 'todas':
            query &= ((db.transacciones.moneda_origen == moneda_filtro) | 
                     (db.transacciones.moneda_destino == moneda_filtro))
        
        transacciones = db(query).select()
        
        # Estadísticas generales
        compras = [t for t in transacciones if t.tipo_operacion == 'compra']
        ventas = [t for t in transacciones if t.tipo_operacion == 'venta']
        
        # Volúmenes por moneda - Compras
        volumen_compras_ves = sum([float(t.monto_origen) for t in compras if t.moneda_origen == 'VES'])
        volumen_compras_usd = sum([float(t.monto_destino) for t in compras if t.moneda_destino == 'USD'])
        volumen_compras_usdt = sum([float(t.monto_destino) for t in compras if t.moneda_destino == 'USDT'])
        volumen_compras_eur = sum([float(t.monto_destino) for t in compras if t.moneda_destino == 'EUR'])
        
        # Volúmenes por moneda - Ventas
        volumen_ventas_usd = sum([float(t.monto_origen) for t in ventas if t.moneda_origen == 'USD'])
        volumen_ventas_usdt = sum([float(t.monto_origen) for t in ventas if t.moneda_origen == 'USDT'])
        volumen_ventas_eur = sum([float(t.monto_origen) for t in ventas if t.moneda_origen == 'EUR'])
        volumen_ventas_ves = sum([float(t.monto_destino) for t in ventas if t.moneda_destino == 'VES'])
        
        # Obtener tasas actuales para conversión consolidada
        tasas_actuales = obtener_tasas_actuales()
        
        # Calcular totales consolidados en VES
        total_ves = volumen_compras_ves + volumen_ventas_ves
        total_usd = volumen_compras_usd + volumen_ventas_usd
        total_usdt = volumen_compras_usdt + volumen_ventas_usdt
        total_eur = volumen_compras_eur + volumen_ventas_eur
        
        total_consolidado_ves = total_ves + \
                               (total_usd * tasas_actuales.get('USD', 0)) + \
                               (total_usdt * tasas_actuales.get('USDT', 0)) + \
                               (total_eur * tasas_actuales.get('EUR', 0))
        
        # Clientes activos (considerando cuentas de las transacciones)
        cuentas_activas = set()
        for t in transacciones:
            if t.cuenta_id:
                cuentas_activas.add(t.cuenta_id)
        
        # Obtener clientes únicos de las cuentas activas
        clientes_activos_set = set()
        for cuenta_id in cuentas_activas:
            cuenta = db(db.cuentas.id == cuenta_id).select().first()
            if cuenta:
                clientes_activos_set.add(cuenta.cliente_id)
        
        clientes_activos = len(clientes_activos_set)
        
        return {
            'mes': fecha.strftime('%B %Y'),
            'fecha_inicio': primer_dia.strftime('%Y-%m-%d'),
            'fecha_fin': ultimo_dia.strftime('%Y-%m-%d'),
            'moneda_filtro': moneda_filtro,
            'total_transacciones': len(transacciones),
            'total_compras': len(compras),
            'total_ventas': len(ventas),
            'volumen_compras_ves': volumen_compras_ves,
            'volumen_compras_usd': volumen_compras_usd,
            'volumen_compras_usdt': volumen_compras_usdt,
            'volumen_compras_eur': volumen_compras_eur,
            'volumen_ventas_usd': volumen_ventas_usd,
            'volumen_ventas_usdt': volumen_ventas_usdt,
            'volumen_ventas_eur': volumen_ventas_eur,
            'volumen_ventas_ves': volumen_ventas_ves,
            'total_ves': total_ves,
            'total_usd': total_usd,
            'total_usdt': total_usdt,
            'total_eur': total_eur,
            'total_consolidado_ves': round(total_consolidado_ves, 2),
            'tasas_conversion': tasas_actuales,
            'total_comisiones': sum([float(t.comision) for t in transacciones]),
            'clientes_activos': clientes_activos,
            'cuentas_activas': len(cuentas_activas)
        }
        
    except Exception as e:
        logger.error(f"Error generando reporte mensual: {str(e)}")
        return {'error': str(e)}

def obtener_tasas_actuales():
    """
    Obtiene las tasas de cambio actuales para conversión consolidada
    Requisito: 8.3
    """
    try:
        # Obtener las tasas más recientes de la tabla tasas_cambio
        tasas = {}
        for moneda in ['USD', 'USDT', 'EUR']:
            tasa_record = db(db.tasas_cambio.moneda == moneda).select(
                orderby=~db.tasas_cambio.fecha_actualizacion,
                limitby=(0, 1)
            ).first()
            if tasa_record:
                tasas[moneda] = float(tasa_record.tasa_compra)
            else:
                tasas[moneda] = 0
        return tasas
    except Exception as e:
        logger.error(f"Error obteniendo tasas actuales: {str(e)}")
        return {'USD': 0, 'USDT': 0, 'EUR': 0}

def obtener_estadisticas_generales():
    """
    Obtiene estadísticas generales para el dashboard
    """
    try:
        # Estadísticas del día actual
        fecha_hoy = datetime.date.today()
        fecha_inicio_hoy = datetime.datetime.combine(fecha_hoy, datetime.time.min)
        
        transacciones_hoy = db(
            db.transacciones.fecha_transaccion >= fecha_inicio_hoy
        ).count()
        
        # Estadísticas del mes actual
        primer_dia_mes = fecha_hoy.replace(day=1)
        fecha_inicio_mes = datetime.datetime.combine(primer_dia_mes, datetime.time.min)
        
        transacciones_mes = db(
            db.transacciones.fecha_transaccion >= fecha_inicio_mes
        ).count()
        
        # Comisiones del mes
        transacciones_mes_data = db(
            db.transacciones.fecha_transaccion >= fecha_inicio_mes
        ).select()
        
        comisiones_mes = sum([float(t.comision) for t in transacciones_mes_data])
        
        # Clientes totales
        total_clientes = db(db.clientes.id > 0).count()
        
        # Cuentas activas
        cuentas_activas = db(db.cuentas.estado == 'activa').count()
        
        return {
            'transacciones_hoy': transacciones_hoy,
            'transacciones_mes': transacciones_mes,
            'comisiones_mes': comisiones_mes,
            'total_clientes': total_clientes,
            'cuentas_activas': cuentas_activas,
            'fecha_consulta': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
    except Exception as e:
        logger.error(f"Error obteniendo estadísticas generales: {str(e)}")
        return {
            'error': str(e),
            'transacciones_hoy': 0,
            'transacciones_mes': 0,
            'comisiones_mes': 0,
            'total_clientes': 0,
            'cuentas_activas': 0
        }

@auth.requires_login()
def exportar_pdf():
    """
    Crear función de exportación a PDF usando ReportLab
    Requisitos: 6.4, 7.5
    """
    try:
        # Importar ReportLab (debe estar instalado)
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
        except ImportError:
            response.flash = "ReportLab no está instalado. Instale con: pip install reportlab"
            redirect(URL('reportes', 'index'))
        
        # Obtener parámetros
        tipo_reporte = request.vars.tipo_reporte or 'transacciones'
        fecha_desde = request.vars.fecha_desde
        fecha_hasta = request.vars.fecha_hasta
        
        # Verificar permisos
        es_admin = auth.has_membership('administrador')
        
        if tipo_reporte == 'transacciones':
            return exportar_transacciones_pdf(fecha_desde, fecha_hasta, es_admin)
        elif tipo_reporte == 'reporte_diario' and es_admin:
            return exportar_reporte_diario_pdf(fecha_desde or datetime.date.today().strftime('%Y-%m-%d'))
        else:
            response.flash = "Tipo de reporte no válido o sin permisos"
            redirect(URL('reportes', 'index'))
            
    except Exception as e:
        logger.error(f"Error en exportar_pdf: {str(e)}")
        response.flash = f"Error exportando PDF: {str(e)}"
        redirect(URL('reportes', 'index'))

def exportar_transacciones_pdf(fecha_desde, fecha_hasta, es_admin):
    """
    Exporta historial de transacciones a PDF
    """
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        
        # Crear buffer en memoria
        buffer = BytesIO()
        
        # Crear documento PDF
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72,
                               topMargin=72, bottomMargin=18)
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Contenido del PDF
        story = []
        
        # Título
        if es_admin:
            titulo = "Reporte de Transacciones - Administrativo"
        else:
            titulo = "Historial de Transacciones - Cliente"
        
        story.append(Paragraph(titulo, title_style))
        story.append(Spacer(1, 12))
        
        # Información del reporte
        info_reporte = f"Generado el: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        if fecha_desde:
            info_reporte += f"<br/>Desde: {fecha_desde}"
        if fecha_hasta:
            info_reporte += f"<br/>Hasta: {fecha_hasta}"
        
        story.append(Paragraph(info_reporte, styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Obtener transacciones
        query = db.transacciones.id > 0
        
        if not es_admin:
            # Para clientes, filtrar por sus transacciones
            cliente = db(db.clientes.user_id == auth.user.id).select().first()
            if not cliente:
                raise Exception("Cliente no encontrado")
            query &= (db.transacciones.cuenta_id == db.cuentas.id) & \
                    (db.cuentas.cliente_id == cliente.id)
        
        if fecha_desde:
            fecha_desde_dt = datetime.datetime.strptime(fecha_desde, '%Y-%m-%d')
            query &= db.transacciones.fecha_transaccion >= fecha_desde_dt
        
        if fecha_hasta:
            fecha_hasta_dt = datetime.datetime.strptime(fecha_hasta, '%Y-%m-%d')
            fecha_hasta_dt = fecha_hasta_dt.replace(hour=23, minute=59, second=59)
            query &= db.transacciones.fecha_transaccion <= fecha_hasta_dt
        
        if es_admin:
            transacciones = db(query).select(
                db.transacciones.ALL,
                db.cuentas.numero_cuenta,
                db.clientes.cedula,
                join=[
                    db.cuentas.on(db.transacciones.cuenta_id == db.cuentas.id),
                    db.clientes.on(db.cuentas.cliente_id == db.clientes.id)
                ],
                orderby=~db.transacciones.fecha_transaccion,
                limitby=(0, 500)  # Máximo 500 registros en PDF
            )
        else:
            transacciones = db(query).select(
                db.transacciones.ALL,
                db.cuentas.numero_cuenta,
                join=db.cuentas.on(db.transacciones.cuenta_id == db.cuentas.id),
                orderby=~db.transacciones.fecha_transaccion,
                limitby=(0, 500)
            )
        
        # Crear tabla de transacciones con información de cuentas por moneda
        if es_admin:
            headers = ['Fecha', 'Comprobante', 'Cliente', 'Cta Origen', 'Cta Destino', 'Tipo', 'Origen', 'Destino', 'Tasa']
        else:
            headers = ['Fecha', 'Comprobante', 'Cta Origen', 'Cta Destino', 'Tipo', 'Origen', 'Destino', 'Tasa']
        
        data = [headers]
        
        for t in transacciones:
            # Obtener información de cuentas origen y destino
            cuenta_origen_str, cuenta_destino_str = obtener_cuentas_transaccion(t.transacciones)
            
            if es_admin:
                fila = [
                    t.transacciones.fecha_transaccion.strftime('%d/%m/%Y %H:%M'),
                    t.transacciones.numero_comprobante,
                    t.clientes.cedula,
                    cuenta_origen_str,
                    cuenta_destino_str,
                    t.transacciones.tipo_operacion.title(),
                    f"{float(t.transacciones.monto_origen):.2f} {t.transacciones.moneda_origen}",
                    f"{float(t.transacciones.monto_destino):.2f} {t.transacciones.moneda_destino}",
                    f"{float(t.transacciones.tasa_aplicada):.4f}"
                ]
            else:
                fila = [
                    t.transacciones.fecha_transaccion.strftime('%d/%m/%Y %H:%M'),
                    t.transacciones.numero_comprobante,
                    cuenta_origen_str,
                    cuenta_destino_str,
                    t.transacciones.tipo_operacion.title(),
                    f"{float(t.transacciones.monto_origen):.2f} {t.transacciones.moneda_origen}",
                    f"{float(t.transacciones.monto_destino):.2f} {t.transacciones.moneda_destino}",
                    f"{float(t.transacciones.tasa_aplicada):.4f}"
                ]
            data.append(fila)
        
        # Crear tabla
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        
        # Resumen
        if transacciones:
            story.append(Spacer(1, 12))
            total_transacciones = len(transacciones)
            total_comisiones = sum([float(t.transacciones.comision) for t in transacciones])
            
            resumen = f"Total de transacciones: {total_transacciones}<br/>"
            resumen += f"Total comisiones: {total_comisiones:.2f} VES"
            
            story.append(Paragraph(resumen, styles['Normal']))
        
        # Generar PDF
        doc.build(story)
        
        # Preparar respuesta
        buffer.seek(0)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename="transacciones_{datetime.date.today().strftime("%Y%m%d")}.pdf"'
        
        return buffer.getvalue()
        
    except Exception as e:
        logger.error(f"Error exportando transacciones a PDF: {str(e)}")
        response.flash = f"Error: {str(e)}"
        redirect(URL('reportes', 'index'))

def exportar_reporte_diario_pdf(fecha_str):
    """
    Exporta reporte diario administrativo a PDF
    """
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        
        # Generar reporte diario
        reporte = generar_reporte_diario(fecha_str)
        
        if 'error' in reporte:
            raise Exception(reporte['error'])
        
        # Crear buffer en memoria
        buffer = BytesIO()
        
        # Crear documento PDF
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72,
                               topMargin=72, bottomMargin=18)
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Contenido del PDF
        story = []
        
        # Título
        story.append(Paragraph(f"Reporte Diario - {reporte['fecha']}", title_style))
        story.append(Spacer(1, 12))
        
        # Resumen ejecutivo
        resumen = f"""
        <b>Resumen Ejecutivo</b><br/>
        Total de transacciones: {reporte['total_transacciones']}<br/>
        Operaciones de compra: {reporte['total_compras']}<br/>
        Operaciones de venta: {reporte['total_ventas']}<br/>
        <br/>
        <b>Compras por Moneda:</b><br/>
        Volumen compras VES: {reporte['volumen_compras_ves']:,.2f}<br/>
        Volumen compras USD: {reporte['volumen_compras_usd']:,.2f}<br/>
        Volumen compras USDT: {reporte['volumen_compras_usdt']:,.2f}<br/>
        Volumen compras EUR: {reporte['volumen_compras_eur']:,.2f}<br/>
        <br/>
        <b>Ventas por Moneda:</b><br/>
        Volumen ventas VES: {reporte.get('volumen_ventas_ves', 0):,.2f}<br/>
        Volumen ventas USD: {reporte['volumen_ventas_usd']:,.2f}<br/>
        Volumen ventas USDT: {reporte['volumen_ventas_usdt']:,.2f}<br/>
        Volumen ventas EUR: {reporte['volumen_ventas_eur']:,.2f}<br/>
        <br/>
        <b>Tasas Promedio:</b><br/>
        Tasa USD promedio: {reporte['tasa_usd_promedio']}<br/>
        Tasa USDT promedio: {reporte['tasa_usdt_promedio']}<br/>
        Tasa EUR promedio: {reporte['tasa_eur_promedio']}<br/>
        <br/>
        Total comisiones: {reporte['total_comisiones']:,.2f} VES<br/>
        <b>Total Consolidado VES: {reporte.get('total_consolidado_ves', 0):,.2f}</b><br/>
        """
        
        story.append(Paragraph(resumen, styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Detalle de transacciones
        if reporte['transacciones_detalle']:
            story.append(Paragraph("<b>Detalle de Transacciones</b>", styles['Heading2']))
            story.append(Spacer(1, 6))
            
            headers = ['Hora', 'Comprobante', 'Tipo', 'Cta Origen', 'Cta Destino', 'Origen', 'Destino', 'Tasa']
            data = [headers]
            
            for t in reporte['transacciones_detalle']:
                # Obtener información de cuentas usando la función auxiliar
                # Crear un objeto temporal con los campos necesarios
                from gluon.storage import Storage
                trans_obj = Storage(
                    cuenta_id=t.get('cuenta_id'),
                    tipo_operacion=t.get('tipo'),
                    moneda_origen=t.get('moneda_origen'),
                    moneda_destino=t.get('moneda_destino')
                )
                cuenta_origen_str, cuenta_destino_str = obtener_cuentas_transaccion(trans_obj)
                
                fila = [
                    t['fecha'],
                    t['comprobante'],
                    t['tipo'].title(),
                    cuenta_origen_str,
                    cuenta_destino_str,
                    f"{t['monto_origen']:.2f} {t['moneda_origen']}",
                    f"{t['monto_destino']:.2f} {t['moneda_destino']}",
                    f"{t['tasa']:.4f}"
                ]
                data.append(fila)
            
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
        
        # Generar PDF
        doc.build(story)
        
        # Preparar respuesta
        buffer.seek(0)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename="reporte_diario_{fecha_str}.pdf"'
        
        return buffer.getvalue()
        
    except Exception as e:
        logger.error(f"Error exportando reporte diario a PDF: {str(e)}")
        response.flash = f"Error: {str(e)}"
        redirect(URL('reportes', 'index'))

@auth.requires_membership('administrador')
def exportar_excel():
    """
    Desarrollar exportación a Excel para reportes administrativos
    Requisitos: 6.4, 7.5
    """
    try:
        # Importar openpyxl (debe estar instalado)
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            response.flash = "openpyxl no está instalado. Instale con: pip install openpyxl"
            redirect(URL('reportes', 'index'))
        
        # Obtener parámetros
        tipo_reporte = request.vars.tipo_reporte or 'transacciones'
        fecha_desde = request.vars.fecha_desde
        fecha_hasta = request.vars.fecha_hasta
        
        if tipo_reporte == 'transacciones':
            return exportar_transacciones_excel(fecha_desde, fecha_hasta)
        elif tipo_reporte == 'reporte_diario':
            return exportar_reporte_diario_excel(fecha_desde or datetime.date.today().strftime('%Y-%m-%d'))
        else:
            response.flash = "Tipo de reporte no válido"
            redirect(URL('reportes', 'index'))
            
    except Exception as e:
        logger.error(f"Error en exportar_excel: {str(e)}")
        response.flash = f"Error exportando Excel: {str(e)}"
        redirect(URL('reportes', 'index'))

def exportar_transacciones_excel(fecha_desde, fecha_hasta):
    """
    Exporta transacciones a Excel con soporte para cuentas por moneda
    Requisito: 8.4
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Transacciones"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        # Encabezados actualizados con información de cuentas por moneda
        headers = ['Fecha', 'Comprobante', 'Cliente', 'Cuenta Origen', 'Moneda Cuenta Origen',
                  'Cuenta Destino', 'Moneda Cuenta Destino', 'Tipo', 'Moneda Origen', 
                  'Monto Origen', 'Moneda Destino', 'Monto Destino', 'Tasa', 'Comisión', 'Estado']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Obtener transacciones con información de cuentas origen y destino
        query = db.transacciones.id > 0
        
        if fecha_desde:
            fecha_desde_dt = datetime.datetime.strptime(fecha_desde, '%Y-%m-%d')
            query &= db.transacciones.fecha_transaccion >= fecha_desde_dt
        
        if fecha_hasta:
            fecha_hasta_dt = datetime.datetime.strptime(fecha_hasta, '%Y-%m-%d')
            fecha_hasta_dt = fecha_hasta_dt.replace(hour=23, minute=59, second=59)
            query &= db.transacciones.fecha_transaccion <= fecha_hasta_dt
        
        transacciones = db(query).select(
            orderby=~db.transacciones.fecha_transaccion
        )
        
        # Llenar datos
        for row, t in enumerate(transacciones, 2):
            # Obtener información de cuentas usando la función auxiliar
            cuenta_origen_str, cuenta_destino_str = obtener_cuentas_transaccion(t)
            
            # Obtener información del cliente desde la cuenta de la transacción
            cliente = None
            cliente_nombre = "N/A"
            cuenta = db(db.cuentas.id == t.cuenta_id).select().first() if t.cuenta_id else None
            if cuenta:
                cliente = db(db.clientes.id == cuenta.cliente_id).select().first()
                if cliente:
                    usuario = db(db.auth_user.id == cliente.user_id).select().first()
                    if usuario:
                        cliente_nombre = f"{usuario.first_name} {usuario.last_name}"
            
            ws.cell(row=row, column=1, value=t.fecha_transaccion.strftime('%d/%m/%Y %H:%M:%S'))
            ws.cell(row=row, column=2, value=t.numero_comprobante)
            ws.cell(row=row, column=3, value=cliente_nombre)
            ws.cell(row=row, column=4, value=cuenta_origen_str)
            ws.cell(row=row, column=5, value=t.moneda_origen)
            ws.cell(row=row, column=6, value=cuenta_destino_str)
            ws.cell(row=row, column=7, value=cuenta_destino.moneda if cuenta_destino else "N/A")
            ws.cell(row=row, column=8, value=t.tipo_operacion.title())
            ws.cell(row=row, column=9, value=t.moneda_origen)
            ws.cell(row=row, column=10, value=float(t.monto_origen))
            ws.cell(row=row, column=11, value=t.moneda_destino)
            ws.cell(row=row, column=12, value=float(t.monto_destino))
            ws.cell(row=row, column=13, value=float(t.tasa_aplicada))
            ws.cell(row=row, column=14, value=float(t.comision))
            ws.cell(row=row, column=15, value=t.estado.title())
        
        # Crear hojas adicionales por tipo de moneda
        for moneda in ['VES', 'USD', 'USDT', 'EUR']:
            ws_moneda = wb.create_sheet(f"Transacciones {moneda}")
            
            # Encabezados
            for col, header in enumerate(headers, 1):
                cell = ws_moneda.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # Filtrar transacciones por moneda
            transacciones_moneda = [t for t in transacciones if t.moneda_origen == moneda or t.moneda_destino == moneda]
            
            # Llenar datos
            for row_idx, t in enumerate(transacciones_moneda, 2):
                cuenta_origen_str, cuenta_destino_str = obtener_cuentas_transaccion(t)
                
                # Obtener información del cliente
                cliente = None
                cliente_nombre = "N/A"
                cuenta = db(db.cuentas.id == t.cuenta_id).select().first() if t.cuenta_id else None
                if cuenta:
                    cliente = db(db.clientes.id == cuenta.cliente_id).select().first()
                    if cliente:
                        usuario = db(db.auth_user.id == cliente.user_id).select().first()
                        if usuario:
                            cliente_nombre = f"{usuario.first_name} {usuario.last_name}"
                
                ws_moneda.cell(row=row_idx, column=1, value=t.fecha_transaccion.strftime('%d/%m/%Y %H:%M:%S'))
                ws_moneda.cell(row=row_idx, column=2, value=t.numero_comprobante)
                ws_moneda.cell(row=row_idx, column=3, value=cliente_nombre)
                ws_moneda.cell(row=row_idx, column=4, value=cuenta_origen_str)
                ws_moneda.cell(row=row_idx, column=5, value=t.moneda_origen)
                ws_moneda.cell(row=row_idx, column=6, value=cuenta_destino_str)
                ws_moneda.cell(row=row_idx, column=7, value=t.moneda_destino)
                ws_moneda.cell(row=row_idx, column=8, value=t.tipo_operacion.title())
                ws_moneda.cell(row=row_idx, column=9, value=t.moneda_origen)
                ws_moneda.cell(row=row_idx, column=10, value=float(t.monto_origen))
                ws_moneda.cell(row=row_idx, column=11, value=t.moneda_destino)
                ws_moneda.cell(row=row_idx, column=12, value=float(t.monto_destino))
                ws_moneda.cell(row=row_idx, column=13, value=float(t.tasa_aplicada))
                ws_moneda.cell(row=row_idx, column=14, value=float(t.comision))
                ws_moneda.cell(row=row_idx, column=15, value=t.estado.title())
        
        # Ajustar ancho de columnas en todas las hojas
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Preparar respuesta
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="transacciones_{datetime.date.today().strftime("%Y%m%d")}.xlsx"'
        
        return buffer.getvalue()
        
    except Exception as e:
        logger.error(f"Error exportando transacciones a Excel: {str(e)}")
        response.flash = f"Error: {str(e)}"
        redirect(URL('reportes', 'index'))

def exportar_reporte_diario_excel(fecha_str):
    """
    Exporta reporte diario a Excel con soporte para cuentas por moneda
    Requisito: 8.4
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        # Generar reporte diario
        reporte = generar_reporte_diario(fecha_str, 'todas')
        
        if 'error' in reporte:
            raise Exception(reporte['error'])
        
        # Crear workbook
        wb = Workbook()
        
        # Hoja de resumen
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        
        # Estilos
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        
        # Título
        ws_resumen.cell(row=1, column=1, value=f"Reporte Diario - {reporte['fecha']}").font = title_font
        
        # Resumen con información por moneda
        resumen_data = [
            ['Total Transacciones', reporte['total_transacciones']],
            ['Total Compras', reporte['total_compras']],
            ['Total Ventas', reporte['total_ventas']],
            ['', ''],
            ['COMPRAS POR MONEDA', ''],
            ['Volumen Compras VES', reporte['volumen_compras_ves']],
            ['Volumen Compras USD', reporte['volumen_compras_usd']],
            ['Volumen Compras USDT', reporte['volumen_compras_usdt']],
            ['Volumen Compras EUR', reporte['volumen_compras_eur']],
            ['', ''],
            ['VENTAS POR MONEDA', ''],
            ['Volumen Ventas VES', reporte.get('volumen_ventas_ves', 0)],
            ['Volumen Ventas USD', reporte['volumen_ventas_usd']],
            ['Volumen Ventas USDT', reporte['volumen_ventas_usdt']],
            ['Volumen Ventas EUR', reporte['volumen_ventas_eur']],
            ['', ''],
            ['TASAS PROMEDIO', ''],
            ['Tasa USD Promedio', reporte['tasa_usd_promedio']],
            ['Tasa USDT Promedio', reporte['tasa_usdt_promedio']],
            ['Tasa EUR Promedio', reporte['tasa_eur_promedio']],
            ['', ''],
            ['Total Comisiones', reporte['total_comisiones']],
            ['Total Consolidado VES', reporte.get('total_consolidado_ves', 0)]
        ]
        
        for row, (concepto, valor) in enumerate(resumen_data, 3):
            ws_resumen.cell(row=row, column=1, value=concepto).font = header_font
            ws_resumen.cell(row=row, column=2, value=valor)
        
        # Hoja de detalle con información de cuentas
        if reporte['transacciones_detalle']:
            ws_detalle = wb.create_sheet("Detalle Transacciones")
            
            # Encabezados actualizados
            headers = ['Hora', 'Comprobante', 'Tipo', 'Cuenta Origen', 'Cuenta Destino',
                      'Moneda Origen', 'Monto Origen', 'Moneda Destino', 'Monto Destino', 
                      'Tasa', 'Comisión']
            
            for col, header in enumerate(headers, 1):
                ws_detalle.cell(row=1, column=col, value=header).font = header_font
            
            # Datos
            for row, t in enumerate(reporte['transacciones_detalle'], 2):
                # Obtener números de cuenta
                cuenta_origen_num = "N/A"
                cuenta_destino_num = "N/A"
                if t.get('cuenta_id'):
                    cuenta = db(db.cuentas.id == t['cuenta_id']).select().first()
                    if cuenta:
                        if t['tipo'] == 'compra':
                            # En compra: cuenta_id es destino, origen es VES
                            cuenta_destino_num = cuenta.numero_cuenta
                            cuenta_ves = db((db.cuentas.cliente_id == cuenta.cliente_id) & 
                                           (db.cuentas.moneda == 'VES')).select().first()
                            cuenta_origen_num = cuenta_ves.numero_cuenta if cuenta_ves else "N/A"
                        else:  # venta
                            # En venta: cuenta_id es origen, destino es VES
                            cuenta_origen_num = cuenta.numero_cuenta
                            cuenta_ves = db((db.cuentas.cliente_id == cuenta.cliente_id) & 
                                           (db.cuentas.moneda == 'VES')).select().first()
                            cuenta_destino_num = cuenta_ves.numero_cuenta if cuenta_ves else "N/A"
                
                ws_detalle.cell(row=row, column=1, value=t['fecha'])
                ws_detalle.cell(row=row, column=2, value=t['comprobante'])
                ws_detalle.cell(row=row, column=3, value=t['tipo'].title())
                ws_detalle.cell(row=row, column=4, value=cuenta_origen_num)
                ws_detalle.cell(row=row, column=5, value=cuenta_destino_num)
                ws_detalle.cell(row=row, column=6, value=t['moneda_origen'])
                ws_detalle.cell(row=row, column=7, value=t['monto_origen'])
                ws_detalle.cell(row=row, column=8, value=t['moneda_destino'])
                ws_detalle.cell(row=row, column=9, value=t['monto_destino'])
                ws_detalle.cell(row=row, column=10, value=t['tasa'])
                ws_detalle.cell(row=row, column=11, value=t['comision'])
        
        # Crear hojas separadas por moneda
        for moneda in ['VES', 'USD', 'USDT', 'EUR']:
            transacciones_moneda = [t for t in reporte['transacciones_detalle'] 
                                   if t['moneda_origen'] == moneda or t['moneda_destino'] == moneda]
            
            if transacciones_moneda:
                ws_moneda = wb.create_sheet(f"Transacciones {moneda}")
                
                # Encabezados
                headers = ['Hora', 'Comprobante', 'Tipo', 'Cuenta Origen', 'Cuenta Destino',
                          'Moneda Origen', 'Monto Origen', 'Moneda Destino', 'Monto Destino', 
                          'Tasa', 'Comisión']
                
                for col, header in enumerate(headers, 1):
                    ws_moneda.cell(row=1, column=col, value=header).font = header_font
                
                # Datos
                for row, t in enumerate(transacciones_moneda, 2):
                    cuenta_origen_num = "N/A"
                    cuenta_destino_num = "N/A"
                    if t.get('cuenta_id'):
                        cuenta = db(db.cuentas.id == t['cuenta_id']).select().first()
                        if cuenta:
                            if t['tipo'] == 'compra':
                                # En compra: cuenta_id es destino, origen es VES
                                cuenta_destino_num = cuenta.numero_cuenta
                                cuenta_ves = db((db.cuentas.cliente_id == cuenta.cliente_id) & 
                                               (db.cuentas.moneda == 'VES')).select().first()
                                cuenta_origen_num = cuenta_ves.numero_cuenta if cuenta_ves else "N/A"
                            else:  # venta
                                # En venta: cuenta_id es origen, destino es VES
                                cuenta_origen_num = cuenta.numero_cuenta
                                cuenta_ves = db((db.cuentas.cliente_id == cuenta.cliente_id) & 
                                               (db.cuentas.moneda == 'VES')).select().first()
                                cuenta_destino_num = cuenta_ves.numero_cuenta if cuenta_ves else "N/A"
                    
                    ws_moneda.cell(row=row, column=1, value=t['fecha'])
                    ws_moneda.cell(row=row, column=2, value=t['comprobante'])
                    ws_moneda.cell(row=row, column=3, value=t['tipo'].title())
                    ws_moneda.cell(row=row, column=4, value=cuenta_origen_num)
                    ws_moneda.cell(row=row, column=5, value=cuenta_destino_num)
                    ws_moneda.cell(row=row, column=6, value=t['moneda_origen'])
                    ws_moneda.cell(row=row, column=7, value=t['monto_origen'])
                    ws_moneda.cell(row=row, column=8, value=t['moneda_destino'])
                    ws_moneda.cell(row=row, column=9, value=t['monto_destino'])
                    ws_moneda.cell(row=row, column=10, value=t['tasa'])
                    ws_moneda.cell(row=row, column=11, value=t['comision'])
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Preparar respuesta
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="reporte_diario_{fecha_str}.xlsx"'
        
        return buffer.getvalue()
        
    except Exception as e:
        logger.error(f"Error exportando reporte diario a Excel: {str(e)}")
        response.flash = f"Error: {str(e)}"
        redirect(URL('reportes', 'index'))